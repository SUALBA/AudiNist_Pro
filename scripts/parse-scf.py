"""
parse-scf.py
Parses the Secure Controls Framework (SCF) Excel spreadsheet and outputs a
lean JSON file used by the AuditNIST Pro Template Library feature.

Usage:
    pip install openpyxl
    python scripts/parse-scf.py

Output:
    data/scf-controls.json
"""

import json
import os
import re
import openpyxl

SHEET_NAME = "SCF 2025.4"
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "..", "secure-controls-framework-scf-2025-4.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "..", "data", "scf-controls.json")

# Fixed column indices (1-based) identified from the SCF 2025.4 sheet structure
COL_SCF_CONTROL = 2    # SCF Control (name)
COL_SCF_ID      = 3    # SCF # (e.g. GOV-01)
COL_SCF_DESC    = 4    # SCF Control Description
COL_SCF_QUESTION = 12  # SCF Control Question
# Possible Solutions columns (firm size classes) — used as implementation guidance
COL_SOL_START   = 7
COL_SOL_END     = 11
# Framework mapping columns
COL_CIS         = 28   # CIS CSC 8.1
COL_COBIT       = 32   # COBIT 2019
COL_ISO27001    = 48   # ISO 27001 2022
COL_NIST_CSF    = 93   # NIST CSF 2.0

# Industry vertical inference: map partial header strings → vertical tag
VERTICAL_KEYWORDS = {
    "healthcare": ["hipaa", "hitech", "cms mars", "hl7"],
    "finance":    ["pci dss", "pci-dss", "pcidss", "sox ", "glba", "finra", "swift"],
    "government": ["fedramp", "nist 800-53", "nist sp 800", "fisma", "disa stig", "cmmc"],
    "defense":    ["cmmc", "dfars", "itar", "nispom", "dod "],
    "privacy":    ["gdpr", "ccpa", "cpra", "pipeda", "lgpd", "appi", "dpdpa"],
}


def clean(val):
    """Return stripped string or empty string for None/non-string values."""
    if val is None:
        return ""
    return str(val).strip()


def parse_mapping(val):
    """Split a comma/semicolon-separated mapping string into a list of IDs."""
    s = clean(val)
    if not s:
        return []
    # Split on common delimiters, strip whitespace
    parts = re.split(r"[,;\n]+", s)
    return [p.strip() for p in parts if p.strip()]


def main():
    print(f"Loading {EXCEL_FILE} ...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    # ── Step 1: Read header row, discover regulatory columns for verticals ──
    headers = {}          # col_index (1-based) → header string
    vertical_cols = {}    # vertical_name → list of col indices

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for col_idx, cell_val in enumerate(first_row, start=1):
        h = clean(cell_val).lower()
        if h:
            headers[col_idx] = h

    # Map headers to verticals
    for col_idx, h in headers.items():
        for vertical, keywords in VERTICAL_KEYWORDS.items():
            if any(kw in h for kw in keywords):
                vertical_cols.setdefault(vertical, []).append(col_idx)

    print("Vertical columns found:")
    for v, cols in vertical_cols.items():
        print(f"  {v}: {len(cols)} columns (e.g. col {cols[0]}: '{headers.get(cols[0], '?')}')")

    # ── Step 2: Parse all data rows ──
    controls = []
    skipped  = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        scf_id   = clean(row[COL_SCF_ID - 1])
        scf_name = clean(row[COL_SCF_CONTROL - 1])
        question = clean(row[COL_SCF_QUESTION - 1])

        if not scf_id or not scf_name:
            skipped += 1
            continue

        # Implementation guidance: concatenate non-empty Possible Solutions columns
        guidance_parts = []
        for c in range(COL_SOL_START - 1, COL_SOL_END):
            v = clean(row[c]) if c < len(row) else ""
            if v:
                guidance_parts.append(v)
        guidance = " | ".join(guidance_parts)

        # Framework mappings
        mappings = {
            "nist-csf": parse_mapping(row[COL_NIST_CSF - 1] if COL_NIST_CSF - 1 < len(row) else None),
            "iso27001":  parse_mapping(row[COL_ISO27001 - 1] if COL_ISO27001 - 1 < len(row) else None),
            "cobit":     parse_mapping(row[COL_COBIT - 1]    if COL_COBIT - 1    < len(row) else None),
            "cis":       parse_mapping(row[COL_CIS - 1]      if COL_CIS - 1      < len(row) else None),
        }

        # Skip controls with no framework mappings (not relevant to our 4 frameworks)
        if not any(mappings.values()):
            skipped += 1
            continue

        # Infer industry verticals
        verticals = ["general"]
        for vertical, cols in vertical_cols.items():
            for col_idx in cols:
                if col_idx - 1 < len(row) and clean(row[col_idx - 1]):
                    verticals.append(vertical)
                    break  # only need one hit per vertical

        controls.append({
            "id":       scf_id,
            "name":     scf_name,
            "question": question,
            "guidance": guidance,
            "mappings": mappings,
            "verticals": list(dict.fromkeys(verticals)),  # deduplicate, preserve order
        })

    wb.close()

    # ── Step 3: Write output ──
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    output = {"version": "scf-2025.4", "controls": controls}
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = os.path.getsize(OUTPUT_FILE) / 1024
    print(f"\nDone.")
    print(f"  Controls included : {len(controls)}")
    print(f"  Controls skipped  : {skipped}")
    print(f"  Output file       : {OUTPUT_FILE}")
    print(f"  File size         : {size_kb:.1f} KB")

    # Breakdown by vertical
    print("\nControls per vertical:")
    for v in ["general", "healthcare", "finance", "government", "defense", "privacy"]:
        count = sum(1 for c in controls if v in c["verticals"])
        print(f"  {v:12s}: {count}")


if __name__ == "__main__":
    main()
